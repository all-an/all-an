from re import A
from openpyxl import load_workbook

file_path = '\test.xlsx'

wb = load_workbook(file_path)

ws = wb['SHEET_NAME']  # or wb.active

ws['A1'] = 123

wb.save(file_path)

print(ws.active_cell)